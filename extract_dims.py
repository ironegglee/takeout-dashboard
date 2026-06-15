import openpyxl, json, codecs

path = r'D:\工作\杂项\外卖指标.xlsx'
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

def extract_cols(ws, col_map, skip_header=True):
    """col_map: {col_index(1-based): field_name}"""
    results = {k: set() for k in col_map.values()}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if skip_header and i == 0:
            continue
        for col_idx, field_name in col_map.items():
            val = row[col_idx - 1] if len(row) >= col_idx else None
            if val is not None and str(val).strip():
                results[field_name].add(str(val).strip())
    return {k: sorted(list(v)) for k, v in results.items()}

# --- 美团数据源 ---
ws_mt = wb['美团数据源']
# 先看表头
print("=== 美团数据源 表头 ===")
for i, row in enumerate(ws_mt.iter_rows(min_row=1, max_row=1, values_only=True)):
    for j, v in enumerate(row):
        print(f"  Col{j+1}: {v}")

# 重新打开因为read_only迭代后无法重置
wb2 = openpyxl.load_workbook(path, data_only=True, read_only=True)
ws_mt2 = wb2['美团数据源']

mt_col_map = {
    5: 'market',      # 市场
    6: 'region_mgr',  # 大区经理  
    7: 'area_mgr',    # 区经理
    4: 'city',        # 城市
    2: 'store_name',  # 门店名称
    21: 'brand',      # 品牌（如果有的话）
}
mt_dims = extract_cols(ws_mt2, mt_col_map)
print("\n=== 美团数据源 维度值 ===")
for k, v in mt_dims.items():
    print(f"  {k} ({len(v)}): {v[:20]}{' ...' if len(v)>20 else ''}")

wb2.close()

# --- 出餐数据源 ---
wb3 = openpyxl.load_workbook(path, data_only=True, read_only=True)
ws_mp = wb3['出餐数据源']
print("\n=== 出餐数据源 表头 ===")
for i, row in enumerate(ws_mp.iter_rows(min_row=1, max_row=1, values_only=True)):
    for j, v in enumerate(row):
        print(f"  Col{j+1}: {v}")

wb3.close()

wb4 = openpyxl.load_workbook(path, data_only=True, read_only=True)
ws_mp2 = wb4['出餐数据源']

mp_col_map = {
    3: 'market',      # 市场
    12: 'region_mgr', # 大区经理
    13: 'area_mgr',   # 区经理
    5: 'city',        # 城市
    2: 'store_name',  # 门店名称
}
mp_dims = extract_cols(ws_mp2, mp_col_map)
print("\n=== 出餐数据源 维度值 ===")
for k, v in mp_dims.items():
    print(f"  {k} ({len(v)}): {v[:20]}{' ...' if len(v)>20 else ''}")

wb4.close()

# 保存结果
output = {
    'mt': mt_dims,
    'mp': mp_dims,
}
with codecs.open(r'C:\Users\CYYS\WorkBuddy\2026-06-01-16-48-48\dimensions.json', 'w', 'utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)
print("\n已保存到 dimensions.json")
