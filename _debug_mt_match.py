import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

OLD_PATH = 'data_sources/OLD.xlsx'
NEW_PATH = 'data_sources/NEW.xlsx'

def parse_date_str(val):
    s = str(val).strip()
    if s == 'nan' or s == 'None' or not s:
        return None
    if len(s) == 8 and s.isdigit():
        return f'{s[:4]}-{s[4:6]}-{s[6:8]}'
    try:
        if hasattr(val, 'strftime'):
            return val.strftime('%Y-%m-%d')
    except:
        pass
    return None

# Read NEW header
wb = openpyxl.load_workbook(NEW_PATH, read_only=True, data_only=True)
ws = wb['美团指标数据源']
header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
wb.close()

header_idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

print('Header idx:')
for k in ['门店编码', '品牌', '门店名称', '日期', '门店id']:
    print(f'  {k}: {header_idx.get(k)}')

# Read 3 rows from NEW
wb = openpyxl.load_workbook(NEW_PATH, read_only=True, data_only=True)
ws = wb['美团指标数据源']
rows = list(ws.iter_rows(min_row=2, max_row=4, values_only=True))
wb.close()

for i, row in enumerate(rows):
    code = str(row[header_idx.get('门店编码', 0)]).strip() if '门店编码' in header_idx and row[header_idx['门店编码']] is not None else ''
    brand = str(row[header_idx.get('品牌', 0)]).strip() if '品牌' in header_idx and row[header_idx['品牌']] is not None else ''
    name_arch = str(row[header_idx.get('门店名称', 0)]).strip() if '门店名称' in header_idx and row[header_idx['门店名称']] is not None else ''
    date_str = parse_date_str(row[header_idx.get('日期', 0)])
    mtid = str(row[header_idx.get('门店id', 0)]).replace('.0','').strip() if '门店id' in header_idx and row[header_idx['门店id']] is not None else ''
    print(f'\nRow {i+2}:')
    print(f'  code={code!r}, brand={brand!r}, name_arch={name_arch!r}, date={date_str}, mtid={mtid}')

# Check mt_arch building from all rows
wb = openpyxl.load_workbook(NEW_PATH, read_only=True, data_only=True)
ws = wb['美团指标数据源']
mt_arch = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    code = str(row[header_idx.get('门店编码', 0)]).strip() if '门店编码' in header_idx and row[header_idx['门店编码']] is not None else ''
    if not code:
        continue
    brand = str(row[header_idx.get('品牌', 0)]).strip() if '品牌' in header_idx and row[header_idx['品牌']] is not None else ''
    name_arch = str(row[header_idx.get('门店名称', 0)]).strip() if '门店名称' in header_idx and row[header_idx['门店名称']] is not None else ''
    if code not in mt_arch:
        mt_arch[code] = {'brand': brand, 'name': name_arch}
wb.close()

print(f'\nmt_arch count: {len(mt_arch)}')

# Check recent_mt building
cutoff_str = '2026-06-29'
recent_mt = {}
wb = openpyxl.load_workbook(NEW_PATH, read_only=True, data_only=True)
ws = wb['美团指标数据源']
for row in ws.iter_rows(min_row=2, values_only=True):
    date_str = parse_date_str(row[header_idx.get('日期', 0)])
    if not date_str or date_str < cutoff_str:
        continue
    mtid = str(row[header_idx.get('门店id', 0)]).replace('.0','').strip() if '门店id' in header_idx and row[header_idx['门店id']] is not None else ''
    if not mtid:
        continue
    code = str(row[header_idx.get('门店编码', 0)]).strip() if '门店编码' in header_idx and row[header_idx['门店编码']] is not None else ''
    if not code:
        continue
    if code not in recent_mt or date_str > recent_mt[code]['date']:
        recent_mt[code] = {'date': date_str, 'mtid': mtid}
wb.close()

print(f'recent_mt count: {len(recent_mt)}')

# Check overlap
matched = 0
for code in recent_mt:
    if code in mt_arch:
        matched += 1
    else:
        if matched < 5:
            print(f'  Unmatched code: {code!r}')

print(f'Matched: {matched}/{len(recent_mt)}')

# Check a specific code
sample_code = list(recent_mt.keys())[0]
print(f'\nSample code: {sample_code!r}')
print(f'  in mt_arch: {sample_code in mt_arch}')
if sample_code in mt_arch:
    print(f'  mt_arch[{sample_code}]: {mt_arch[sample_code]}')
