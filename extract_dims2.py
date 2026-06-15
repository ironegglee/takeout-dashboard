# -*- coding: utf-8 -*-
import openpyxl, json, sys

path = r'D:\工作\杂项\外卖指标.xlsx'
OUT = r'C:\Users\CYYS\WorkBuddy\2026-06-01-16-48-48\dimensions.json'

def extract_unique(ws, col_indices, skip_header=True):
    """col_indices: list of (1-based col_index, field_name)"""
    sets = {name: set() for _, name in col_indices}
    row_count = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if skip_header and i == 0:
            continue
        for col_idx, name in col_indices:
            if len(row) >= col_idx and row[col_idx-1] is not None:
                v = str(row[col_idx-1]).strip()
                if v:
                    sets[name].add(v)
        row_count += 1
        if row_count % 10000 == 0:
            pass  # progress
    return {name: sorted(list(s)) for name, s in sets.items()}, row_count

result = {}

# === 美团数据源 ===
wb1 = openpyxl.load_workbook(path, data_only=True, read_only=True)
ws_mt = wb1['美团数据源']
mt_cols = [
    (2, 'store_name'),
    (4, 'city'),
    (5, 'market'),
    (6, 'region_mgr'),
    (7, 'area_mgr'),
]
mt_dims, mt_rows = extract_unique(ws_mt, mt_cols)
wb1.close()
result['mt'] = mt_dims
result['mt_rows'] = mt_rows

# === 出餐数据源 ===
wb2 = openpyxl.load_workbook(path, data_only=True, read_only=True)
ws_mp = wb2['出餐数据源']
mp_cols = [
    (2, 'store_name'),
    (3, 'market'),
    (5, 'city'),
    (12, 'region_mgr'),
    (13, 'area_mgr'),
]
mp_dims, mp_rows = extract_unique(ws_mp, mp_cols)
wb2.close()
result['mp'] = mp_dims
result['mp_rows'] = mp_rows

with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print('SUCCESS', file=sys.stderr)
