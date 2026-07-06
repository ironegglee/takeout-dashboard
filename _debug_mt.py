import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

OLD_PATH = 'data_sources/OLD.xlsx'
NEW_PATH = 'data_sources/NEW.xlsx'

# Read NEW header
wb = openpyxl.load_workbook(NEW_PATH, read_only=True, data_only=False)
ws = wb['美团指标数据源']
header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
wb.close()

print('Header cols:', len(header))
print('First 10:', header[:10])
print('Brand col:', header.index('品牌') if '品牌' in header else 'NOT FOUND')
print('Store code col:', header.index('门店编码') if '门店编码' in header else 'NOT FOUND')
print('Store name.1 col:', header.index('门店名称.1') if '门店名称.1' in header else 'NOT FOUND')
print('Date col:', header.index('日期') if '日期' in header else 'NOT FOUND')
print('Store id col:', header.index('门店id') if '门店id' in header else 'NOT FOUND')

# Read first 3 data rows
wb = openpyxl.load_workbook(NEW_PATH, read_only=True, data_only=False)
ws = wb['美团指标数据源']
rows = list(ws.iter_rows(min_row=2, max_row=4, values_only=True))
wb.close()

for i, row in enumerate(rows):
    print(f'\nRow {i+2}:')
    for j, h in enumerate(header[:15]):
        print(f'  {h}: {row[j]}')
