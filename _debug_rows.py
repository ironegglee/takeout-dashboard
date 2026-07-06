import openpyxl, sys
from datetime import datetime, timedelta
sys.stdout.reconfigure(encoding='utf-8')

OLD_PATH = 'data_sources/OLD.xlsx'
NEW_PATH = 'data_sources/NEW.xlsx'

def parse_date_str(val):
    s = str(val).strip()
    if s == 'nan' or s == 'None' or not s:
        return None
    if len(s) == 8 and s.isdigit():
        return f'{s[:4]}-{s[4:6]}-{s[6:8]}'
    try:
        if hasattr(val, 'strftime'):
            return val.strftime('%Y-%m-%d')
    except:
        pass
    return None

# Read NEW header
wb = openpyxl.load_workbook(NEW_PATH, read_only=True, data_only=True)
ws = wb['美团指标数据源']
header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
wb.close()

header_idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

# Read OLD header
wb = openpyxl.load_workbook(OLD_PATH, read_only=True, data_only=True)
ws = wb['美团指标数据源']
old_header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
wb.close()

old_header_idx = {str(h).strip(): i for i, h in enumerate(old_header) if h is not None}

print('NEW header:', header[:50])
print('OLD header:', old_header[:50])
print('Equal:', header == old_header)

# Read first 2 rows from OLD
wb = openpyxl.load_workbook(OLD_PATH, read_only=True, data_only=True)
ws = wb['美团指标数据源']
rows = list(ws.iter_rows(min_row=2, max_row=3, values_only=True))
wb.close()

for i, row in enumerate(rows):
    print(f'\nOLD Row {i+2}:')
    for k, idx in [('门店编码', 40), ('品牌', 43), ('门店名称', 41), ('日期', 0)]:
        val = row[idx] if idx < len(row) else 'OUT OF RANGE'
        print(f'  {k}[{idx}]: {val!r}')

# Read first 2 rows from NEW
wb = openpyxl.load_workbook(NEW_PATH, read_only=True, data_only=True)
ws = wb['美团指标数据源']
rows = list(ws.iter_rows(min_row=2, max_row=3, values_only=True))
wb.close()

for i, row in enumerate(rows):
    print(f'\nNEW Row {i+2}:')
    for k, idx in [('门店编码', 40), ('品牌', 43), ('门店名称', 41), ('日期', 0)]:
        val = row[idx] if idx < len(row) else 'OUT OF RANGE'
        print(f'  {k}[{idx}]: {val!r}')
