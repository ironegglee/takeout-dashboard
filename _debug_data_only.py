import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

OLD_PATH = 'data_sources/OLD.xlsx'

wb = openpyxl.load_workbook(OLD_PATH, read_only=True, data_only=True)
ws = wb['美团指标数据源']
header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))

header_idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

print('With data_only=True:')
rows = list(ws.iter_rows(min_row=2, max_row=6, values_only=True))
wb.close()

for i, row in enumerate(rows):
    code = str(row[header_idx.get('门店编码', 0)]).strip() if '门店编码' in header_idx and row[header_idx['门店编码']] is not None else ''
    brand = str(row[header_idx.get('品牌', 0)]).strip() if '品牌' in header_idx and row[header_idx['品牌']] is not None else ''
    name_arch = str(row[header_idx.get('门店名称', 0)]).strip() if '门店名称' in header_idx and row[header_idx['门店名称']] is not None else ''
    print(f'Row {i+2}: code={code}, brand={brand}, name_arch={name_arch}')
