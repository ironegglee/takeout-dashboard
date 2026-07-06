import openpyxl, sys
from datetime import datetime, timedelta
sys.stdout.reconfigure(encoding='utf-8')

OLD_PATH = 'data_sources/OLD.xlsx'
NEW_PATH = 'data_sources/NEW.xlsx'

def parse_date_str(val):
    s = str(val).strip()
    if s == 'nan' or s == 'None' or not s: return None
    if len(s) == 8 and s.isdigit(): return f'{s[:4]}-{s[4:6]}-{s[6:8]}'
    try:
        if hasattr(val, 'strftime'): return val.strftime('%Y-%m-%d')
    except: pass
    return None

BRAND_MAP = {'茶颜': '茶颜悦色', '茶颜悦色': '茶颜悦色', '鸳央': '鸳央咖啡', '鸳央咖啡': '鸳央咖啡', '墨柠': '墨柠', '古德墨柠': '墨柠', '昼夜': '昼夜诗', '昼夜诗': '昼夜诗'}

def normalize_brand(val):
    v = str(val).strip() if val is not None else ''
    if not v or v == 'nan' or v == 'None': return ''
    if v in BRAND_MAP: return BRAND_MAP[v]
    if '茶颜' in v: return '茶颜悦色'
    if '鸳央' in v or '鸯央' in v: return '鸳央咖啡'
    if '墨柠' in v or '古德' in v: return '墨柠'
    if '昼夜' in v: return '昼夜诗'
    if '饼坊' in v or '饼行' in v: return '茶颜悦色'
    return v

# Read header
wb = openpyxl.load_workbook(NEW_PATH, read_only=True, data_only=True)
ws = wb['美团指标数据源']
header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
wb.close()

header_idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

# Read all rows from OLD + NEW, deduplicate by (date, mtid)
mt_data = {}
for path in [OLD_PATH, NEW_PATH]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['美团指标数据源']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 44: continue
        date_str = parse_date_str(row[header_idx.get('日期', 0)])
        mtid = str(row[header_idx.get('门店id', 2)]).replace('.0','').strip() if row[header_idx.get('门店id', 2)] is not None else ''
        if date_str and mtid:
            mt_data[(date_str, mtid)] = row
    wb.close()

print(f'Total MT records: {len(mt_data)}')

# Get latest date
all_dates = sorted(set(k[0] for k in mt_data.keys()))
latest_mt_date = all_dates[-1] if all_dates else '2026-07-01'
latest_dt = datetime.strptime(latest_mt_date, '%Y-%m-%d')
cutoff_dt = latest_dt - timedelta(days=6)
cutoff_str = cutoff_dt.strftime('%Y-%m-%d')
print(f'Latest date: {latest_mt_date}, cutoff: {cutoff_str}')

# Build mt_arch
mt_arch = {}
for row in mt_data.values():
    code = str(row[header_idx.get('门店编码', 40)]).strip() if row[header_idx.get('门店编码', 40)] is not None else ''
    if not code: continue
    if code not in mt_arch:
        brand = normalize_brand(row[header_idx.get('品牌', 43)])
        name = str(row[header_idx.get('门店名称', 41)]).strip() if row[header_idx.get('门店名称', 41)] is not None else ''
        mt_arch[code] = {'brand': brand, 'name': name}

print(f'mt_arch: {len(mt_arch)} codes')

# Build recent_mt
recent_mt = {}
for (date_str, mtid), row in mt_data.items():
    if date_str >= cutoff_str:
        code = str(row[header_idx.get('门店编码', 40)]).strip() if row[header_idx.get('门店编码', 40)] is not None else ''
        if not code: continue
        if code not in recent_mt or date_str > recent_mt[code]['date']:
            recent_mt[code] = {'date': date_str, 'row': row}

print(f'recent_mt: {len(recent_mt)} codes')

# Build mt_stores
mt_stores = []
mt_unmatched = 0
for code, info in recent_mt.items():
    row = info['row']
    arch_info = mt_arch.get(code)
    if not arch_info:
        mt_unmatched += 1
        continue
    
    shop_score = float(row[header_idx.get('店铺分', 7)]) if row[header_idx.get('店铺分', 7)] is not None else 0
    
    mt_stores.append({
        'code': code,
        'name': arch_info['name'],
        'brand': arch_info['brand'],
        'shop_score': shop_score,
    })

print(f'Before filter: {len(mt_stores)}')
mt_stores = [s for s in mt_stores if s.get('name') and s.get('brand') and s['brand'] != 'nan']
print(f'After filter: {len(mt_stores)}')
print(f'Unmatched: {mt_unmatched}')

if mt_stores:
    print(f'First store: {mt_stores[0]}')
    print(f'Last store: {mt_stores[-1]}')
