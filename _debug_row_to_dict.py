import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

NEW_PATH = 'data_sources/NEW.xlsx'

wb = openpyxl.load_workbook(NEW_PATH, read_only=True, data_only=True)
ws = wb['美团指标数据源']
header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
wb.close()

header_idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

# Same as gen_data_v4.py
mt_header_idx = header_idx

def row_to_dict(row, header_idx):
    return {k: (row[v] if v < len(row) else None) for k, v in header_idx.items()}

wb = openpyxl.load_workbook(NEW_PATH, read_only=True, data_only=True)
ws = wb['美团指标数据源']
rows = list(ws.iter_rows(min_row=2, max_row=3, values_only=True))
wb.close()

for i, row in enumerate(rows):
    d = row_to_dict(row, mt_header_idx)
    print(f'\nRow {i+2}:')
    print(f'  d.get("门店编码") = {d.get("门店编码")!r}')
    print(f'  d.get("门店名称") = {d.get("门店名称")!r}')
    print(f'  d.get("品牌") = {d.get("品牌")!r}')
    print(f'  d.get("日期") = {d.get("日期")!r}')
    print(f'  d.get("门店id") = {d.get("门店id")!r}')
    print(f'  len(d) = {len(d)}')
    print(f'  len(row) = {len(row)}')

# Test with OLD
wb = openpyxl.load_workbook('data_sources/OLD.xlsx', read_only=True, data_only=True)
ws = wb['美团指标数据源']
row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
wb.close()

d = row_to_dict(row, mt_header_idx)
print(f'\nOLD Row 2:')
print(f'  d.get("门店编码") = {d.get("门店编码")!r}')
print(f'  d.get("门店名称") = {d.get("门店名称")!r}')
print(f'  d.get("品牌") = {d.get("品牌")!r}')
print(f'  len(d) = {len(d)}')
print(f'  len(row) = {len(row)}')
