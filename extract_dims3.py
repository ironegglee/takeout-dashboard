# -*- coding: utf-8 -*-
import openpyxl, json, sys

path = r'D:\工作\杂项\外卖指标.xlsx'
OUT = r'C:\Users\CYYS\WorkBuddy\2026-06-01-16-48-48\dims_simple.json'

result = {}

# === 美团数据源 ===
wb1 = openpyxl.load_workbook(path, data_only=True, read_only=True)
ws_mt = wb1['美团数据源']

# 先打印表头和前三行数据
header_mt = list(next(ws_mt.iter_rows(min_row=1, max_row=1, values_only=True)))
result['mt_header'] = [str(v) for v in header_mt]

# 收集每列的样本值
col_samples = {}
for i, row in enumerate(ws_mt.iter_rows(min_row=2, values_only=True)):
    if i >= 5: break
    for j, v in enumerate(row):
        if v is not None:
            sv = str(v).strip()
            if sv:
                col_samples.setdefault(j, set()).add(sv)

result['mt_col_samples'] = {str(j): list(s)[:3] for j, s in col_samples.items()}

# 提取关键维度 (尝试找到正确的列)
# 根据之前分析: Col5=市场, Col6=省份, Col7=城市, Col8=大区, Col9=市场名, Col10=区域
# 让我用表头名称来匹配
market_idx = region_mgr_idx = area_mgr_idx = city_idx = store_idx = None
for j, h in enumerate(header_mt):
    h_str = str(h).strip() if h else ''
    if '市场' in h_str and j > 4:  # 跳过前面的"市场"（可能是门店所在市场等）
        market_idx = j
    if '大区' in h_str or h_str == '大区':
        region_mgr_idx = j
    if '区域' in h_str:
        area_mgr_idx = j
    if '城市' in h_str:
        city_idx = j
    if '门店名称' in h_str:
        store_idx = j

result['mt_col_mapping'] = {
    'market': market_idx,
    'region_mgr': region_mgr_idx,
    'area_mgr': area_mgr_idx,
    'city': city_idx,
    'store_name': store_idx,
}

# 提取唯一值
dims = {'market': set(), 'region_mgr': set(), 'area_mgr': set(), 'city': set()}
idx_map = {'market': market_idx, 'region_mgr': region_mgr_idx, 'area_mgr': area_mgr_idx, 'city': city_idx}
for row in ws_mt.iter_rows(min_row=2, values_only=True):
    for name, idx in idx_map.items():
        if idx is not None and len(row) > idx:
            v = row[idx]
            if v is not None:
                sv = str(v).strip()
                if sv and len(sv) < 50:  # 排除长ID
                    dims[name].add(sv)

result['mt_dims'] = {k: sorted(list(v)) for k, v in dims.items()}

wb1.close()

# === 出餐数据源 ===
wb2 = openpyxl.load_workbook(path, data_only=True, read_only=True)
ws_mp = wb2['出餐数据源']

header_mp = list(next(ws_mp.iter_rows(min_row=1, max_row=1, values_only=True)))
result['mp_header'] = [str(v) for v in header_mp]

# 收集每列的样本值
col_samples2 = {}
for i, row in enumerate(ws_mp.iter_rows(min_row=2, values_only=True)):
    if i >= 5: break
    for j, v in enumerate(row):
        if v is not None:
            sv = str(v).strip()
            if sv:
                col_samples2.setdefault(j, set()).add(sv)

result['mp_col_samples'] = {str(j): list(s)[:3] for j, s in col_samples2.items()}

# 维度
mp_market_idx = mp_region_mgr_idx = mp_area_mgr_idx = mp_city_idx = mp_store_idx = None
for j, h in enumerate(header_mp):
    h_str = str(h).strip() if h else ''
    if '市场' in h_str:
        mp_market_idx = j
    if '大区' in h_str:
        mp_region_mgr_idx = j
    if '区域' in h_str:
        mp_area_mgr_idx = j
    if '城市' in h_str:
        mp_city_idx = j
    if '门店' in h_str and '名称' not in h_str:
        # 门店代码/名称等
        pass
    if '门店名称' in h_str:
        mp_store_idx = j

result['mp_col_mapping'] = {
    'market': mp_market_idx,
    'region_mgr': mp_region_mgr_idx,
    'area_mgr': mp_area_mgr_idx,
    'city': mp_city_idx,
    'store_name': mp_store_idx,
}

mp_dims = {'market': set(), 'region_mgr': set(), 'area_mgr': set(), 'city': set()}
mp_idx_map = {'market': mp_market_idx, 'region_mgr': mp_region_mgr_idx, 'area_mgr': mp_area_mgr_idx, 'city': mp_city_idx}
for row in ws_mp.iter_rows(min_row=2, values_only=True):
    for name, idx in mp_idx_map.items():
        if idx is not None and len(row) > idx:
            v = row[idx]
            if v is not None:
                sv = str(v).strip()
                if sv and len(sv) < 50:
                    mp_dims[name].add(sv)

result['mp_dims'] = {k: sorted(list(v)) for k, v in mp_dims.items()}

wb2.close()

with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print('DONE', file=sys.stderr)
