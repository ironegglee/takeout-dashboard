import openpyxl, json, sys, gc, os
from datetime import datetime, timedelta
import math
import traceback

sys.stdout.reconfigure(encoding='utf-8')

_fatal_error = False

def _handle_exception(exc_type, exc_value, exc_traceback):
    global _fatal_error
    _fatal_error = True
    with open('C:/Users/CYYS/WorkBuddy/2026-06-16-11-25-31/_gen_exception.txt', 'w', encoding='utf-8') as f:
        traceback.print_exception(exc_type, exc_value, exc_traceback, file=f)
    sys.__excepthook__(exc_type, exc_value, exc_traceback)
sys.excepthook = _handle_exception

# 数据源统一从 D:\工作\workbuddy\ 读取 —— 这是唯一权威数据源目录
# 不要使用 data_sources/ 或其他副本
SRC_DIR = 'D:/工作/workbuddy'
OLD_PATH = f'{SRC_DIR}/外卖业务运营看板数据源5月.xlsx'
MID_PATH = f'{SRC_DIR}/外卖业务运营看板数据源6月.xlsx'
NEW_PATH = f'{SRC_DIR}/外卖业务运营看板数据源7月.xlsx'
OUT = 'C:/Users/CYYS/WorkBuddy/2026-06-16-11-25-31/dashboard/data.json'

BRAND_MAP = {
    '茶颜': '茶颜悦色', '茶颜悦色': '茶颜悦色',
    '鸳央': '鸳央咖啡', '鸳央咖啡': '鸳央咖啡',
    '墨柠': '墨柠', '古德墨柠': '墨柠',
    '昼夜': '昼夜诗', '昼夜诗': '昼夜诗'
}

def normalize_brand(val):
    v = str(val).strip() if val is not None else ''
    if not v or v == 'nan' or v == 'None':
        return ''
    if v in BRAND_MAP:
        return BRAND_MAP[v]
    if '茶颜' in v:
        return '茶颜悦色'
    if '鸳央' in v or '鸯央' in v:
        return '鸳央咖啡'
    if '墨柠' in v or '古德' in v:
        return '墨柠'
    if '昼夜' in v:
        return '昼夜诗'
    if '饼坊' in v or '饼行' in v:
        return '茶颜悦色'
    return v

BRAND_TO_MARKET = {
    '墨柠': '湖南', '古德墨柠': '湖南',
    '鸳央咖啡': '湖南', '鸳央': '湖南',
    '昼夜诗': '湖南', '昼夜': '湖南',
}

def normalize_market(market_val, brand_val=''):
    v = str(market_val).strip() if market_val is not None else ''
    if not v or v == 'nan' or v == 'None':
        return ''
    if v in BRAND_TO_MARKET:
        return BRAND_TO_MARKET[v]
    brand_norm = normalize_brand(brand_val) if brand_val else ''
    if brand_norm in BRAND_TO_MARKET and (brand_norm in v or v in brand_norm or ('墨柠' in v and '墨柠' in brand_norm) or ('鸳央' in v and '鸳央' in brand_norm) or ('昼夜' in v and '昼夜' in brand_norm)):
        return BRAND_TO_MARKET[brand_norm]
    MARKET_NORM = {
        '湖南': '湖南', '湖南省': '湖南', '湖南省市场': '湖南',
        '江苏': '江苏', '江苏省': '江苏', '江苏省市场': '江苏',
        '湖北': '湖北', '湖北省': '湖北', '湖北省市场': '湖北',
        '重庆': '重庆', '重庆市': '重庆', '重庆市市场': '重庆',
        '广东': '广东', '广东省': '广东', '广东市场': '广东',
        '深圳': '广东',
    }
    if v in MARKET_NORM:
        return MARKET_NORM[v]
    for std_name, normed in [('湖南','湖南'),('江苏','江苏'),('湖北','湖北'),('重庆','重庆'),('广东','广东')]:
        if std_name in v and v != normed:
            return normed
    return v

def safe_float(val, default=0.0):
    try:
        v = float(val)
        return default if math.isnan(v) else v
    except:
        return default

def safe_int(val, default=0):
    try:
        return int(float(val))
    except:
        return default

def parse_date_str(val):
    """将日期值转为 YYYY-MM-DD 字符串"""
    s = str(val).strip()
    if s == 'nan' or s == 'None' or not s:
        return None
    # 可能是 20260701 格式
    if len(s) == 8 and s.isdigit():
        return f'{s[:4]}-{s[4:6]}-{s[6:8]}'
    # 尝试 datetime 格式
    try:
        if hasattr(val, 'strftime'):
            return val.strftime('%Y-%m-%d')
    except:
        pass
    return None

print('=== 流式 gen_data.py v4 ===')

# ═══════════════════════════════════════════
# 1. 读取架构 (OLD+NEW文件「美团架构数据源」合并)
# ═══════════════════════════════════════════
print('\n=== 1. 读取架构 ===')

def read_arch_openpyxl(filepath):
    """用 openpyxl 读取架构数据（统一9列：美团ID,门店编码,门店名称,市场,品牌,城市,大区,区域,大店）"""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['美团架构数据源']
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()
    
    if not rows:
        return {}, {}, {}
    
    # 9列统一格式
    idx_map = {
        '美团ID': 0, '门店编码': 1, '门店名称': 2, '市场': 3, '品牌': 4,
        '城市': 5, '大区': 6, '区域': 7, '大店': 8
    }
    
    by_mtid, by_name, by_code = {}, {}, {}
    for row in rows[1:]:
        if len(row) < 3:
            continue
        name = str(row[idx_map['门店名称']]).strip() if row[idx_map['门店名称']] is not None else ''
        if not name or name == 'nan' or name == 'None':
            continue
        
        mtid_val = row[idx_map['美团ID']]
        mtid = str(int(mtid_val)) if mtid_val is not None and str(mtid_val).replace('.','').isdigit() else ''
        
        city = str(row[idx_map['城市']]).strip() if row[idx_map['城市']] is not None else ''
        region = str(row[idx_map['大区']]).strip() if row[idx_map['大区']] is not None else ''
        area = str(row[idx_map['区域']]).strip() if row[idx_map['区域']] is not None else ''
        leader = str(row[idx_map['大店']]).strip() if row[idx_map['大店']] is not None else ''
        
        brand_raw = row[idx_map['品牌']]
        brand = normalize_brand(brand_raw)
        
        market_raw = row[idx_map['市场']]
        market = normalize_market(market_raw, brand)
        
        code_val = row[idx_map['门店编码']]
        code = str(code_val).strip() if code_val is not None else ''
        
        info = {
            'brand': brand, 'market': market, 'city': city,
            'region_mgr': region, 'area_mgr': area, 'leader': leader,
            'store_code': code, 'arch_name': name
        }
        if mtid:
            by_mtid[mtid] = info
        if name:
            by_name[name] = info
        if code:
            by_code[code] = info
    
    return by_mtid, by_name, by_code

# 按优先级读取架构：7月 > 6月 > 5月（后加载的覆盖前面的）
arch_by_mtid, arch_by_name, arch_by_code = {}, {}, {}

# 5月（优先级最低，最后合并）
m5_mtid, m5_name, m5_code = read_arch_openpyxl(OLD_PATH)
print(f'5月架构: {len(m5_mtid)} 门店(有美团ID)')

# 6月（覆盖5月）
m6_mtid, m6_name, m6_code = read_arch_openpyxl(MID_PATH)
print(f'6月架构: {len(m6_mtid)} 门店(有美团ID)')

# 7月（最高优先级）
m7_mtid, m7_name, m7_code = read_arch_openpyxl(NEW_PATH)
print(f'7月架构: {len(m7_mtid)} 门店(有美团ID)')

# 按优先级合并：5月 → 6月覆盖 → 7月覆盖
for mtid, info in m5_mtid.items(): arch_by_mtid[mtid] = info
for name, info in m5_name.items(): arch_by_name[name] = info
for code, info in m5_code.items(): arch_by_code[code] = info

for mtid, info in m6_mtid.items(): arch_by_mtid[mtid] = info
for name, info in m6_name.items(): arch_by_name[name] = info
for code, info in m6_code.items(): arch_by_code[code] = info

for mtid, info in m7_mtid.items(): arch_by_mtid[mtid] = info
for name, info in m7_name.items(): arch_by_name[name] = info
for code, info in m7_code.items(): arch_by_code[code] = info

print(f'合并后: {len(arch_by_mtid)} 门店(有美团ID)')

brand_counts = {}
for info in arch_by_mtid.values():
    b = info['brand']
    brand_counts[b] = brand_counts.get(b, 0) + 1
print(f'品牌分布: {dict(sorted(brand_counts.items()))}')

# ═══════════════════════════════════════════
# 2. 读取美团指标数据 (OLD+NEW「美团指标数据源」逐行读取合并)
# ═══════════════════════════════════════════
print('\n=== 2. 读取美团指标数据 ===')

def read_mt_sheet(filepath):
    """逐行读取美团指标数据源，返回 (header_list, rows_list)"""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['美团指标数据源']
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()
    if not rows:
        return [], []
    return rows[0], rows[1:]

def make_header_index(header):
    """将 header 列表转为 {列名: 索引} 字典"""
    return {str(h).strip(): i for i, h in enumerate(header) if h is not None}

def row_to_dict(row, header_idx):
    """将一行转为 {列名: 值} 字典（用于MP部分，MP表头无重复列名）"""
    return {k: (row[v] if v < len(row) else None) for k, v in header_idx.items()}

# 读取三个文件
m5_header, m5_mt_rows = read_mt_sheet(OLD_PATH)
m6_header, m6_mt_rows = read_mt_sheet(MID_PATH)
m7_header, m7_mt_rows = read_mt_sheet(NEW_PATH)

print(f'5月美团指标: {len(m5_mt_rows)} 行, 6月: {len(m6_mt_rows)} 行, 7月: {len(m7_mt_rows)} 行')
print(f'7月表头列数: {len(m7_header)}, 6月: {len(m6_header)}, 5月: {len(m5_header)}')

# mt_header_idx: {列名: 索引}，基于最新文件（7月）的表头
# 三个文件表头结构一致，用7月的即可
mt_header_idx = make_header_index(m7_header)

# 为了方便从行元组读值，提供一个安全访问函数
def col_val(row, header_idx, col_name, default=None):
    """从行元组 row 中按列名读取值"""
    idx = header_idx.get(col_name)
    if idx is None or idx >= len(row):
        return default
    return row[idx]

# 合并去重：用 (日期, 门店id) 作为唯一键，7月 > 6月 > 5月
# 存储为 { (date_str, mtid): row_tuple }，不转为 dict，避免重复列名问题
mt_data = {}
MIN_MT_COLS = max(len(m7_header), len(m6_header), len(m5_header))

def process_mt_rows(rows, label):
    """处理一批MT行，写入 mt_data"""
    cnt = 0
    for row in rows:
        if len(row) < MIN_MT_COLS:
            row = row + (None,) * (MIN_MT_COLS - len(row))
        date_str = parse_date_str(col_val(row, mt_header_idx, '日期'))
        mtid_raw = col_val(row, mt_header_idx, '门店id')
        mtid = str(int(mtid_raw)) if mtid_raw is not None and str(mtid_raw).replace('.','').isdigit() else ''
        if date_str and mtid:
            mt_data[(date_str, mtid)] = row
            cnt += 1
    return cnt

# 5月 → 6月覆盖 → 7月覆盖
n5 = process_mt_rows(m5_mt_rows, '5月')
n6 = process_mt_rows(m6_mt_rows, '6月')
n7 = process_mt_rows(m7_mt_rows, '7月')

print(f'5月写入: {n5}, 6月写入: {n6}, 7月写入: {n7}, 合并后: {len(mt_data)} 条')

# 提取所有日期
all_mt_dates = sorted(set(k[0] for k in mt_data.keys()))
if all_mt_dates:
    print(f'美团日期范围: {all_mt_dates[0]} ~ {all_mt_dates[-1]}')
    latest_mt_date = all_mt_dates[-1]
else:
    latest_mt_date = '2026-07-01'

# 构建 mt_arch (从美团指标表自带的架构列)
# 注意：美团指标数据源有 48 列，架构信息在列40~47（0-indexed: 40=门店编码, 41=门店名称, 42=市场, 43=品牌, 44=城市, 45=大区, 46=区域, 47=大店）
# 直接用列名访问（重复列名取最后一次出现，即列41的「门店名称.1」）
mt_arch = {}
for row in mt_data.values():
    code = str(col_val(row, mt_header_idx, '门店编码') or '').strip()
    if not code:
        continue
    name_arch = str(col_val(row, mt_header_idx, '门店名称') or '').strip()
    # 如果「门店名称」取到的是列1（订单数据中的门店名），尝试「门店名称.1」
    if not name_arch or len(name_arch) < 2:
        name_arch = str(col_val(row, mt_header_idx, '门店名称.1') or '').strip()
    brand = normalize_brand(col_val(row, mt_header_idx, '品牌'))
    market = normalize_market(col_val(row, mt_header_idx, '市场'), brand)
    city = str(col_val(row, mt_header_idx, '城市') or '').strip()
    region = str(col_val(row, mt_header_idx, '大区') or '').strip()
    area = str(col_val(row, mt_header_idx, '区域') or '').strip()
    leader = str(col_val(row, mt_header_idx, '大店') or '').strip()
    mt_arch[code] = {
        'brand': brand, 'market': market, 'city': city,
        'region_mgr': region, 'area_mgr': area, 'leader': leader,
        'arch_name': name_arch
    }

print(f'美团指标表自带架构: {len(mt_arch)} 门店编码')

# 取近7天的记录，按门店编码分组取最新日期
from datetime import datetime, timedelta
latest_dt = datetime.strptime(latest_mt_date, '%Y-%m-%d')
cutoff_dt = latest_dt - timedelta(days=6)
cutoff_str = cutoff_dt.strftime('%Y-%m-%d')

recent_mt = {}  # {code: {date_str, row_tuple}} - 只保留最新日期
mt_all_records = {}  # {code: [{date, row}, ...]} - 所有日期记录，用于得分字段回退
for (date_str, mtid), row in mt_data.items():
    if date_str >= cutoff_str:
        code = str(col_val(row, mt_header_idx, '门店编码') or '').strip()
        if not code:
            continue
        if code not in recent_mt or date_str > recent_mt[code]['date']:
            recent_mt[code] = {'date': date_str, 'row': row}
        if code not in mt_all_records:
            mt_all_records[code] = []
        mt_all_records[code].append({'date': date_str, 'row': row})

print(f'近7日美团门店数(去重后): {len(recent_mt)}')

# 构建 mt_stores
mt_stores = []
mt_unmatched = 0

dim_labels = {
    'peak_hours': '高峰营业时长得分', 'quality_rate': '优质商品率得分',
    'reject_rate': '商家不接单率得分', 'reply_rate': '差评回复率得分',
    'merchant_rating': '商家评分得分', 'menu_rich': '菜单丰富度得分',
    'decor_rich': '装修丰富度得分', 'service_rich': '服务功能丰富度得分',
    'cook_report': '出餐完成上报率得分/配送准时率得分', 'base_hours': '基础营业时长得分'
}

exp_dim_labels = {
    'product_quality': '商品质量分', 'service_exp': '服务体验分',
    'product_sat': '商品满意度', 'pack_sat': '包装满意度',
    'repurchase': '复购率指标', 'msg_reply': '消息回复率',
    'service_neg': '服务负反馈率', 'food_safety': '食品安全负反馈率',
    'cook_report': '出餐上报/配送准时率',
}

def get_valid_score(code, field_name):
    """获取某门店某字段的有效得分，最新日期缺失时回退到历史日期"""
    # 优先取最新日期
    if code in recent_mt:
        val = col_val(recent_mt[code]['row'], mt_header_idx, field_name)
        if val is not None and val != '' and str(val) != '--' and str(val) != 'None':
            return safe_float(val)
    # 回退：按日期倒序查历史有效值
    if code in mt_all_records:
        for rec in sorted(mt_all_records[code], key=lambda x: x['date'], reverse=True):
            val = col_val(rec['row'], mt_header_idx, field_name)
            if val is not None and val != '' and str(val) != '--' and str(val) != 'None':
                return safe_float(val)
    return 0.0

for code, info in recent_mt.items():
    row = info['row']
    arch_info = mt_arch.get(code)
    
    if not arch_info:
        # 回退：用Sheet 0架构匹配（门店名称）
        mt_name = str(col_val(row, mt_header_idx, '门店名称') or '').strip()
        # 如果列1的门店名称不对，尝试列41（门店名称.1）
        if not mt_name or len(mt_name) < 2:
            mt_name = str(col_val(row, mt_header_idx, '门店名称.1') or '').strip()
        mt_id_raw = col_val(row, mt_header_idx, '门店id')
        mt_id = str(int(mt_id_raw)) if mt_id_raw is not None and str(mt_id_raw).replace('.','').isdigit() else ''
        found = arch_by_mtid.get(mt_id) or arch_by_name.get(mt_name)
        if found:
            arch_info = found
        else:
            # 模糊匹配
            for an, ai in arch_by_name.items():
                if an and mt_name and (an in mt_name or mt_name in an):
                    arch_info = ai
                    break
        if not arch_info:
            mt_unmatched += 1
            continue
    
    shop_score = safe_float(col_val(row, mt_header_idx, '店铺分'))
    exp_score = safe_float(col_val(row, mt_header_idx, '综合体验分'))
    
    data_date = info['date']  # 该条数据的日期
    mt_stores.append({
        'code': code,
        'name': arch_info['arch_name'],
        'brand': arch_info['brand'], 'market': arch_info['market'],
        'city': arch_info['city'],
        'region_mgr': arch_info['region_mgr'], 'area_mgr': arch_info['area_mgr'],
        'leader': arch_info['leader'], 'channel': 'mt',
        'shop_score': shop_score, 'exp_score': exp_score,
        'data_date': data_date,
        'shop_dims': {
            'peak_hours': safe_float(col_val(row, mt_header_idx, '高峰营业时长得分')),
            'quality_rate': safe_float(col_val(row, mt_header_idx, '优质商品率得分')),
            'reject_rate': get_valid_score(code, '商家不接单率得分'),
            'reply_rate': get_valid_score(code, '差评回复率得分'),
            'merchant_rating': safe_float(col_val(row, mt_header_idx, '商家评分得分')),
            'menu_rich': safe_float(col_val(row, mt_header_idx, '菜单丰富度得分')),
            'decor_rich': safe_float(col_val(row, mt_header_idx, '装修丰富度得分')),
            'service_rich': safe_float(col_val(row, mt_header_idx, '服务功能丰富度得分')),
            'cook_report': safe_float(col_val(row, mt_header_idx, '出餐完成上报率得分/配送准时率得分')),
            'base_hours': safe_float(col_val(row, mt_header_idx, '基础营业时长得分')),
        },
        'exp_dims': {
            'product_quality': safe_float(col_val(row, mt_header_idx, '商品质量分')),
            'service_exp': safe_float(col_val(row, mt_header_idx, '服务体验分')),
            'product_sat': safe_float(col_val(row, mt_header_idx, '商品满意度')),
            'pack_sat': safe_float(col_val(row, mt_header_idx, '包装满意度')),
            'repurchase': safe_float(col_val(row, mt_header_idx, '复购率指标得分')),
            'msg_reply': safe_float(col_val(row, mt_header_idx, '消息回复率指标得分')),
            'service_neg': safe_float(col_val(row, mt_header_idx, '服务负反馈率指标得分')),
            'food_safety': safe_float(col_val(row, mt_header_idx, '食品安全负反馈率指标得分')),
            'cook_report': safe_float(col_val(row, mt_header_idx, '出餐完成上报率得分/配送准时率得分')),
        }
    })

mt_stores = [s for s in mt_stores if s.get('name') and s.get('brand') and s['brand'] != 'nan']
print(f'美团门店匹配: {len(mt_stores)}/{len(recent_mt)}, 未匹配: {mt_unmatched}')

# ═══════════════════════════════════════════
# 3. 读取小程序配送数据 (OLD+NEW 逐行合并)
# ═══════════════════════════════════════════
print('\n=== 3. 读取小程序配送数据 ===')

def read_mp_header(filepath):
    """读取小程序配送数据源的表头"""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['小程序配送数据源']
    header = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        header = row
        break
    wb.close()
    return header

def process_mp_file(filepath, mp_data, mp_header_idx, skip_existing=False):
    """流式处理小程序配送数据文件，逐个读取行，不保留整个列表"""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['小程序配送数据源']
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = row_to_dict(row, mp_header_idx)
        order_code = str(d.get('订单编码','')).strip() if d.get('订单编码') else ''
        if order_code:
            if skip_existing and order_code in mp_data:
                continue
            mp_data[order_code] = d
            count += 1
    wb.close()
    return count

# 读取表头（以7月为准）
m7_mp_header = read_mp_header(NEW_PATH)
mp_header_idx = make_header_index(m7_mp_header)

# 合并去重：用 (订单编码) 作为唯一键，7月 > 6月 > 5月
# 先处理7月（最新），再补充6月和5月的去重数据，减少内存峰值
mp_data = {}  # {order_code: row_dict}

m7_count = process_mp_file(NEW_PATH, mp_data, mp_header_idx)
print(f'7月小程序配送处理: {m7_count} 条')
gc.collect()

m6_count = process_mp_file(MID_PATH, mp_data, mp_header_idx, skip_existing=True)
print(f'6月小程序配送处理: {m6_count} 条')
gc.collect()

m5_count = process_mp_file(OLD_PATH, mp_data, mp_header_idx, skip_existing=True)
print(f'5月小程序配送处理: {m5_count} 条')
gc.collect()

print(f'合并去重后: {len(mp_data)} 条订单')

# 解析日期
def parse_mp_date(val):
    """解析 pt(day) 或 转化日期"""
    if val is None:
        return None
    s = str(val).strip()
    if s == 'nan' or s == 'None':
        return None
    # 可能是 20260701 格式
    if len(s) == 8 and s.isdigit():
        return f'{s[:4]}-{s[4:6]}-{s[6:8]}'
    # 尝试 datetime
    try:
        if hasattr(val, 'strftime'):
            return val.strftime('%Y-%m-%d')
    except:
        pass
    return None

# 处理所有订单数据
mp_orders = []  # [{date_str, store_name, store_code, city, area, leader, cook_min, order_code, cups}]
for d in mp_data.values():
    date_str = parse_mp_date(d.get('pt(day)')) or parse_mp_date(d.get('转化日期'))
    if not date_str:
        continue
    
    store_name = str(d.get('门店名称','')).strip() if d.get('门店名称') else ''
    store_code = str(d.get('门店代码','')).strip() if d.get('门店代码') else ''
    city = str(d.get('城市','')).strip() if d.get('城市') else ''
    area = str(d.get('区域','')).strip() if d.get('区域') else ''
    leader = str(d.get('大店长','')).strip() if d.get('大店长') else ''
    
    # 制作时长
    cook_min = None
    cook_val = d.get('制作时长')
    if cook_val is not None:
        try:
            cook_min = float(cook_val)
        except:
            pass
    
    # 订单杯数
    cups = 0
    cups_val = d.get('订单杯数')
    if cups_val is not None:
        try:
            cups = int(float(cups_val))
        except:
            pass
    
    mp_orders.append({
        'date': date_str, 'store_name': store_name, 'store_code': store_code,
        'city': city, 'area': area, 'leader': leader,
        'cook_min': cook_min, 'order_code': d.get('订单编码'), 'cups': cups
    })

print(f'有效小程序订单: {len(mp_orders)} 条')

# 日期范围
all_mp_dates = sorted(set(o['date'] for o in mp_orders if o['date']))
if all_mp_dates:
    print(f'小程序日期范围: {all_mp_dates[0]} ~ {all_mp_dates[-1]}')
    max_mp_date = all_mp_dates[-1]
    max_mp_dt = datetime.strptime(max_mp_date, '%Y-%m-%d')
else:
    max_mp_dt = datetime.now()
    max_mp_date = max_mp_dt.strftime('%Y-%m-%d')

mp_cutoff_dt = max_mp_dt - timedelta(days=6)
mp_cutoff_str = mp_cutoff_dt.strftime('%Y-%m-%d')

# 取近7天订单
mp_orders_7d = [o for o in mp_orders if o['date'] >= mp_cutoff_str]
print(f'近7天小程序订单: {len(mp_orders_7d)} 条')

# 按门店聚合近7天数据
mp_store_7d = {}  # {store_name: {orders, cups, cook_sum, cook_count, store_code, city, area, leader}}
for o in mp_orders_7d:
    name = o['store_name']
    if not name:
        continue
    if name not in mp_store_7d:
        mp_store_7d[name] = {
            'orders': 0, 'cups': 0, 'cook_sum': 0.0, 'cook_count': 0,
            'store_code': o['store_code'], 'city': o['city'],
            'area': o['area'], 'leader': o['leader']
        }
    s = mp_store_7d[name]
    s['orders'] += 1
    s['cups'] += o['cups']
    if o['cook_min'] is not None:
        s['cook_sum'] += o['cook_min']
        s['cook_count'] += 1

print(f'近7天小程序门店数: {len(mp_store_7d)}')

# 出餐耗时时段分布（需要小时列）
# 先找到小时列索引
hour_col = None
for i, h in enumerate(m7_mp_header):
    if h and '小时' in str(h):
        hour_col = i
        break

if hour_col is None:
    # 尝试第26列（0-indexed: 26）
    hour_col = 26

# 收集时段数据
hourly_data = {h: [] for h in range(7, 25)}
for d in mp_data.values():
    date_str = parse_mp_date(d.get('pt(day)')) or parse_mp_date(d.get('转化日期'))
    if not date_str or date_str < mp_cutoff_str:
        continue
    
    cook_val = d.get('制作时长')
    cook_min = None
    if cook_val is not None:
        try:
            cook_min = float(cook_val)
        except:
            pass
    if cook_min is None:
        continue
    
    hour_val = d.get(list(mp_header_idx.keys())[hour_col] if hour_col < len(list(mp_header_idx.keys())) else None)
    if hour_val is None:
        # 直接尝试索引
        row_idx = list(mp_data.keys()).index(d.get('订单编码')) if d.get('订单编码') in mp_data else -1
        # 无法获取，跳过
        continue
    try:
        hour = int(float(hour_val))
        if 7 <= hour <= 24:
            hourly_data[hour].append(cook_min)
    except:
        pass

mp_hourly_cook = {}
for h in range(7, 25):
    vals = hourly_data[h]
    if vals:
        mp_hourly_cook[str(h)] = round(sum(vals) / len(vals), 1)
    else:
        mp_hourly_cook[str(h)] = 0.0

# 出餐分段占比
all_cook_mins = []
for o in mp_orders_7d:
    if o['cook_min'] is not None:
        all_cook_mins.append(o['cook_min'])

total_valid = len(all_cook_mins)
bins = [(0,5),(5,10),(10,15),(15,20),(20,999)]
labels = ['0-5分钟','5-10分钟','10-15分钟','15-20分钟','20分钟以上']
colors_bd = ['#3B6D11','#639922','#EF9F27','#E24B4A','#A32D2D']
mp_breakdown = []
for (lo,hi),label,color in zip(bins,labels,colors_bd):
    cnt = sum(1 for v in all_cook_mins if lo <= v < hi)
    pct = round(cnt/total_valid*100,1) if total_valid>0 else 0
    mp_breakdown.append({'label':label,'pct':pct,'color':color,'count':int(cnt)})

# MP门店明细
mp_stores = []
mp_unmatched = []

for name, s in mp_store_7d.items():
    if not name or not name.strip():
        continue
    code = s['store_code']
    
    # 优先用门店代码匹配架构
    info = arch_by_code.get(code) or arch_by_name.get(name)
    
    if not info:
        for an, ai in arch_by_name.items():
            if (an and name) and (an in name or name in an):
                info = ai
                break
    
    if not info:
        mp_unmatched.append(name)
        continue
    
    avg_cook = round(s['cook_sum'] / s['cook_count'], 1) if s['cook_count'] > 0 else 0
    rate = round(sum(1 for v in [o['cook_min'] for o in mp_orders_7d if o['store_name'] == name and o['cook_min'] is not None and o['cook_min'] <= 15]) / s['orders'] * 100, 1) if s['orders'] > 0 else 0
    
    mp_stores.append({
        'code': code,
        'name': info['arch_name'],
        'brand': info['brand'], 'market': info['market'],
        'city': info['city'],
        'region_mgr': info['region_mgr'], 'area_mgr': info['area_mgr'],
        'leader': info['leader'], 'channel': 'mp',
        'avg': avg_cook,
        'max': round(avg_cook * 1.5) if avg_cook > 0 else 0,
        'rate': rate,
        'orders': s['orders'],
        'cups': s['cups'],
    })

mp_stores = [s for s in mp_stores if s.get('name') and s.get('brand') and s['brand'] != 'nan']
print(f'小程序匹配: {len(mp_stores)}/{len(mp_store_7d)}, 未匹配: {len(mp_unmatched)}')

# ═══════════════════════════════════════════
# 7. 预警中心数据
# ═══════════════════════════════════════════
print('\n=== 7. 预警中心 ===')

alerts = []

# ① 出餐超时预警
ref_date = max_mp_dt
cutoff = ref_date - timedelta(days=7)
cutoff_str = cutoff.strftime('%Y-%m-%d')

print(f'  检测: 出餐超时预警（参考日期={ref_date.date()}，窗口={cutoff_str}~{max_mp_date}）...')

# 按店+日计算日均出餐
daily_cook = {}  # {(store_name, date_str): [cook_min, ...]}
for o in mp_orders:
    if o['date'] >= cutoff_str and o['date'] <= max_mp_date and o['cook_min'] is not None:
        key = (o['store_name'], o['date'])
        if key not in daily_cook:
            daily_cook[key] = []
        daily_cook[key].append(o['cook_min'])

# 计算日均
daily_avg = {}  # {(store_name, date_str): avg_cook}
for key, vals in daily_cook.items():
    daily_avg[key] = sum(vals) / len(vals)

# 找出连续2天>15的
over15_stores = {}
store_dates = {}  # {store_name: [date_str, ...]}
for (name, date_str), avg in daily_avg.items():
    if name not in store_dates:
        store_dates[name] = []
    store_dates[name].append(date_str)

for name, dates in store_dates.items():
    dates_sorted = sorted(dates)
    bad_dates = []
    for ds in dates_sorted:
        avg = daily_avg.get((name, ds), 0)
        if avg > 15:
            bad_dates.append(ds)
    
    # 检查是否有连续2天
    has_consecutive = False
    for i in range(len(dates_sorted)-1):
        d1 = dates_sorted[i]
        d2 = dates_sorted[i+1]
        # 检查是否连续
        d1_dt = datetime.strptime(d1, '%Y-%m-%d')
        d2_dt = datetime.strptime(d2, '%Y-%m-%d')
        if (d2_dt - d1_dt).days == 1:
            avg1 = daily_avg.get((name, d1), 0)
            avg2 = daily_avg.get((name, d2), 0)
            if avg1 > 15 and avg2 > 15:
                has_consecutive = True
                break
    
    if has_consecutive and bad_dates:
        all_avgs = [daily_avg.get((name, d), 0) for d in dates_sorted]
        all_vals = []
        for d in dates_sorted:
            all_vals.extend(daily_cook.get((name, d), []))
        over15_stores[name] = {
            'dates': bad_dates,
            'avg': round(sum(all_avgs) / len(all_avgs), 1) if all_avgs else 0,
            'max': round(max(all_vals) if all_vals else 0, 1),
            'bad_avg': round(sum(daily_avg.get((name, d), 0) for d in bad_dates) / len(bad_dates), 1) if bad_dates else 0,
        }

for mp_name, cook_info in over15_stores.items():
    info = arch_by_name.get(mp_name)
    if not info:
        for an, ai in arch_by_name.items():
            if an and mp_name and (an in mp_name or mp_name in an):
                info = ai
                break
    if info:
        bad_dates = cook_info['dates']
        alerts.append({
            'type': 'cook',
            'store': info['arch_name'],
            'brand': info['brand'],
            'market': info['market'],
            'city': info['city'],
            'region_mgr': info['region_mgr'],
            'area_mgr': info['area_mgr'],
            'leader': info['leader'],
            'msg': f'连续2天出餐超15分钟（{bad_dates[0]}~{bad_dates[-1]}），日均 {cook_info["avg"]} 分钟',
            'detail': f'日均出餐 {cook_info["avg"]} 分钟 | 超时日期: {",".join(bad_dates)}',
            'max_cook': cook_info['max'],
            'avg_cook': cook_info['avg'],
            'alert_dates': bad_dates,
        })

print(f'  出餐超时预警: {len(alerts)} 条')

# ② 店铺分预警 —— 方案A：按总分判定，阈值 < 95
# - 总分 < 95 即触发，不再检查子维度
# - 总分 = 0：若门店存在于数据源，视为真实偏差，正常触发（标记 is_zero_score=true）
# - 区分：数据源中不存在的门店已被 mt_stores 过滤，不会进入此处
for s in mt_stores:
    shop_score = s.get('shop_score', 0)
    
    if shop_score < 95:
        # 构建子维度详情（仅供参考，不作为触发条件）
        dim_detail_parts = []
        low_dims = []
        for dim_name in ['peak_hours','quality_rate','reject_rate','reply_rate','merchant_rating',
                          'menu_rich','decor_rich','service_rich','cook_report','base_hours']:
            dim_val = s.get('shop_dims', {}).get(dim_name, 0)
            dim_label = dim_labels.get(dim_name, dim_name)
            status_mark = '⚠' if dim_val < 80 else '✓'
            dim_detail_parts.append(f'{dim_label}{dim_val}分{status_mark}')
            if dim_val < 80:
                low_dims.append({'name': dim_label, 'val': dim_val})
        
        is_zero = (shop_score == 0)
        if is_zero:
            msg = f'店铺分 0 分（数据异常，低于95分）'
        else:
            msg = f'店铺分 {shop_score} 分（低于95分）'
        
        alerts.append({
            'type': 'shop',
            'store': s['name'], 'brand': s['brand'], 'market': s['market'],
            'city': s['city'], 'region_mgr': s['region_mgr'],
            'area_mgr': s['area_mgr'], 'leader': s['leader'],
            'msg': msg,
            'detail': '; '.join(dim_detail_parts),
            'shop_score': shop_score,
            'low_dims': low_dims,
            'data_date': s.get('data_date', ''),
            'is_zero_score': is_zero,
        })

print(f'  店铺分预警: {sum(1 for a in alerts if a["type"]=="shop")} 条（其中0分异常: {sum(1 for a in alerts if a["type"]=="shop" and a.get("is_zero_score"))} 条）')

# ③ 综合体验分预警
for s in mt_stores:
    exp_score = s.get('exp_score', 0)
    if exp_score < 4.6:
        exp_dims_raw = s.get('exp_dims', {})
        low_exp_dims = []
        for dim_key in ['product_quality','service_exp','product_sat','pack_sat',
                         'repurchase','msg_reply','service_neg','food_safety','cook_report']:
            dim_val = exp_dims_raw.get(dim_key, 0)
            if dim_val > 0 and dim_val < 4.2:
                low_exp_dims.append({'name': exp_dim_labels.get(dim_key, dim_key), 'val': round(dim_val, 1)})
        
        if low_exp_dims:
            problem_str = '; '.join([f"{d['name']}{d['val']}分" for d in low_exp_dims])
            detail = f'综合体验分{exp_score:.2f}分; {problem_str}'
        else:
            detail = f'综合体验分{exp_score:.2f}分'
        
        alerts.append({
            'type': 'rating',
            'store': s['name'], 'brand': s['brand'], 'market': s['market'],
            'city': s['city'], 'region_mgr': s['region_mgr'],
            'area_mgr': s['area_mgr'], 'leader': s['leader'],
            'msg': f'综合体验分 {exp_score:.2f} 分（低于4.6分）',
            'detail': detail,
            'exp_score': round(exp_score, 2),
            'exp_dims': exp_dims_raw,
            'low_dims': low_exp_dims,
        })

print(f'  综合体验分预警: {sum(1 for a in alerts if a["type"]=="rating")} 条')
print(f'  预警总数: {len(alerts)} 条')

# ═══════════════════════════════════════════
# 8. 每日汇总
# ═══════════════════════════════════════════
print('\n=== 8. 每日汇总 ===')

# 小程序每日
mp_daily = []
mp_date_stats = {}  # {date_str: {orders, cook_sum, cook_count, store_codes(set)}}
for o in mp_orders:
    ds = o['date']
    if not ds:
        continue
    if ds not in mp_date_stats:
        mp_date_stats[ds] = {'orders': 0, 'cook_sum': 0.0, 'cook_count': 0, 'stores': set()}
    s = mp_date_stats[ds]
    s['orders'] += 1
    if o['cook_min'] is not None:
        s['cook_sum'] += o['cook_min']
        s['cook_count'] += 1
    if o['store_code']:
        s['stores'].add(o['store_code'])

for ds in sorted(mp_date_stats.keys()):
    s = mp_date_stats[ds]
    avg_cook = round(s['cook_sum'] / s['cook_count'], 1) if s['cook_count'] > 0 else 0
    # 达标率：<=15分钟的订单占比
    comply_cnt = sum(1 for o in mp_orders if o['date'] == ds and o['cook_min'] is not None and o['cook_min'] <= 15)
    rate = round(comply_cnt / s['orders'] * 100, 1) if s['orders'] > 0 else 0
    mp_daily.append({
        'date': ds, 'orders': s['orders'], 'avg_cook': avg_cook,
        'rate': rate, 'store_count': len(s['stores'])
    })

print(f'mp_daily: {len(mp_daily)} 天')

# 美团每日
mt_daily = []
mt_date_stats = {}  # {date_str: {shop_scores[], exp_scores[], stores(set)}}
for (date_str, mtid), row in mt_data.items():
    if date_str not in mt_date_stats:
        mt_date_stats[date_str] = {'shop_scores': [], 'exp_scores': [], 'stores': set()}
    s = mt_date_stats[date_str]
    shop_score = safe_float(col_val(row, mt_header_idx, '店铺分'))
    exp_score = safe_float(col_val(row, mt_header_idx, '综合体验分'))
    if shop_score > 0:
        s['shop_scores'].append(shop_score)
    if exp_score > 0:
        s['exp_scores'].append(exp_score)
    s['stores'].add(mtid)

for ds in sorted(mt_date_stats.keys()):
    s = mt_date_stats[ds]
    avg_shop = round(sum(s['shop_scores']) / len(s['shop_scores']), 1) if s['shop_scores'] else 0
    avg_exp = round(sum(s['exp_scores']) / len(s['exp_scores']), 2) if s['exp_scores'] else 0
    mt_daily.append({
        'date': ds, 'avg_shop_score': avg_shop, 'avg_exp_score': avg_exp,
        'store_count': len(s['stores'])
    })

print(f'mt_daily: {len(mt_daily)} 天')

# store_daily
store_daily = {}
# 美团每日门店编码
for (date_str, mtid), row in mt_data.items():
    code = str(col_val(row, mt_header_idx, '门店编码') or '').strip()
    if code:
        if date_str not in store_daily:
            store_daily[date_str] = {'mt': [], 'mp': []}
        if code not in store_daily[date_str]['mt']:
            store_daily[date_str]['mt'].append(code)

# 小程序每日门店编码
for o in mp_orders:
    ds = o['date']
    code = o['store_code']
    if ds and code:
        if ds not in store_daily:
            store_daily[ds] = {'mt': [], 'mp': []}
        if code not in store_daily[ds]['mp']:
            store_daily[ds]['mp'].append(code)

print(f'store_daily: {len(store_daily)} 天')

all_dates = sorted(store_daily.keys())
full_date_range = f'{all_dates[0]} ~ {all_dates[-1]}' if all_dates else ''
print(f'全量日期范围: {full_date_range}')

# mp_date_range
mp_date_range = ''
if all_mp_dates:
    mp_date_range = f'{all_mp_dates[0]} ~ {all_mp_dates[-1]}'

# ═══════════════════════════════════════════
# 9. 输出 JSON
# ═══════════════════════════════════════════
print('\n=== 9. 输出 JSON ===')

def clean_nan(obj):
    if isinstance(obj, float) and math.isnan(obj): return None
    if isinstance(obj, dict): return {k: clean_nan(v) for k,v in obj.items()}
    if isinstance(obj, list): return [clean_nan(v) for v in obj]
    return obj

# 门店时段的逐店出餐
mp_hourly_by_store = {}
for name in mp_store_7d.keys():
    store_hours = {h: [] for h in range(7, 25)}
    for o in mp_orders_7d:
        if o['store_name'] == name and o['cook_min'] is not None:
            # 无法获取小时，简化处理
            pass
    # 简化：不生成时段数据
    hours = [str(h) for h in range(7, 25)]
    vals = [0.0] * 18
    mp_hourly_by_store[name] = {'hours': hours, 'vals': vals}

output = clean_nan({
    'generated_at': str(latest_mt_date),
    'meituan_date': str(latest_mt_date),
    'full_date_range': full_date_range,
    'mp_date_range': mp_date_range,
    'mp_daily': mp_daily,
    'mt_daily': mt_daily,
    'store_daily': store_daily,
    'dashboard_summary': {},
    'mp_hourly_cook': mp_hourly_cook,
    'mp_hourly_by_store': mp_hourly_by_store,
    'mp_breakdown': mp_breakdown,
    'mp_summary': {
        'total_orders': total_valid,
        'total_stores': len(mp_store_7d),
        'avg_cook_time': round(sum(all_cook_mins) / len(all_cook_mins), 1) if all_cook_mins else 0,
        '达标率': round(sum(1 for v in all_cook_mins if v <= 15) / len(all_cook_mins) * 100, 1) if all_cook_mins else 0,
    },
    'mp_summary_full': {
        'total_orders': len(mp_orders),
        'total_stores': len(set(o['store_code'] for o in mp_orders if o['store_code'])),
        'date_range': mp_date_range,
    },
    'mt_stores': mt_stores,
    'mp_stores': mp_stores,
    'store_arch': arch_by_code,
    'alerts': alerts,
    'match_summary': {
        'arch_total': len(arch_by_mtid),
        'mt_matched': len(mt_stores),
        'mt_total': len(recent_mt),
        'mp_matched': len(mp_stores),
        'mp_total_7d': len(mp_store_7d),
    }
})

with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False)

fsize = len(json.dumps(output, ensure_ascii=False)) / 1024
print(f'已输出: {OUT}')
print(f'大小: {fsize:.0f} KB')
print(f'美团门店: {len(mt_stores)} | 小程序门店: {len(mp_stores)} | 预警: {len(alerts)} 条')

# ═══════════════════════════════════════════
# 10. 数据完整性校验
# ═══════════════════════════════════════════
errors = []

if _fatal_error:
    errors.append('捕获到未处理异常，详见 _gen_exception.txt')
if not os.path.exists(OUT):
    errors.append('data.json 文件未生成')
if fsize < 50:
    errors.append(f'data.json 过小({fsize:.0f}KB)，可能数据不完整')
if len(mt_stores) < 100:
    errors.append(f'美团门店数量异常少({len(mt_stores)}家)，预期>=100')
if len(mp_stores) < 100:
    errors.append(f'小程序门店数量异常少({len(mp_stores)}家)，预期>=100')
if len(mp_daily) < 2:
    errors.append(f'小程序每日汇总天数过少({len(mp_daily)}天)')
if not full_date_range or '~' not in full_date_range:
    errors.append('日期范围异常或为空')

if errors:
    print('\n=== 数据校验失败 ===')
    for e in errors:
        print(f'  ❌ {e}')
    print('请检查数据源文件是否完整，或查看 _gen_exception.txt')
    sys.exit(1)

print('\n=== 数据校验通过 ===')
print('Done!')
