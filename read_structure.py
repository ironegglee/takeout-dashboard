import openpyxl, json

path = r'D:\工作\杂项\外卖指标.xlsx'
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

result = {}
for sn in wb.sheetnames:
    ws = wb[sn]
    cols = []
    for cell in ws[1]:  # header row
        cols.append(str(cell.value) if cell.value is not None else '-')
    result[sn] = {"rows": ws.max_row, "cols": ws.max_column, "header": cols}

with open(r'C:\Users\CYYS\WorkBuddy\2026-06-01-16-48-48\xlsx_structure.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print("Done")
