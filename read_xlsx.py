import openpyxl, sys, json

path = r'D:\工作\杂项\外卖指标.xlsx'
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
result = {"sheets": wb.sheetnames, "data": {}}

for sn in wb.sheetnames:
    ws = wb[sn]
    sheet_info = {"rows": ws.max_row, "cols": ws.max_column, "sample": []}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(8, ws.max_row), values_only=True)):
        vals = [str(v)[:60] if v is not None else '-' for v in row]
        sheet_info["sample"].append(vals)
    result["data"][sn] = sheet_info

print(json.dumps(result, ensure_ascii=False, indent=2))
