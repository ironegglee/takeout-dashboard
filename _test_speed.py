import openpyxl, sys, time
sys.stdout.reconfigure(encoding='utf-8')

OLD_PATH = 'data_sources/OLD.xlsx'

# Test data_only=False speed
start = time.time()
wb = openpyxl.load_workbook(OLD_PATH, read_only=True, data_only=False)
ws = wb['美团指标数据源']
count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    count += 1
wb.close()
t1 = time.time() - start
print(f'data_only=False: {count} rows in {t1:.1f}s')

# Test data_only=True speed
start = time.time()
wb = openpyxl.load_workbook(OLD_PATH, read_only=True, data_only=True)
ws = wb['美团指标数据源']
count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    count += 1
wb.close()
t2 = time.time() - start
print(f'data_only=True: {count} rows in {t2:.1f}s')

# Check what ArrayFormula looks like
wb = openpyxl.load_workbook(OLD_PATH, read_only=True, data_only=False)
ws = wb['美团指标数据源']
row = next(ws.iter_rows(min_row=2, max_row=2, values_only=False))
wb.close()

cell = row[40]  # 门店编码 column
print(f'Cell type: {type(cell.value)}')
print(f'Cell value: {cell.value}')
if hasattr(cell.value, 'text'):
    print(f'Formula text: {cell.value.text}')
if hasattr(cell.value, 'ref'):
    print(f'Formula ref: {cell.value.ref}')
