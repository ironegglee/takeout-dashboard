import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

OLD_PATH = 'data_sources/OLD.xlsx'

wb = openpyxl.load_workbook(OLD_PATH, read_only=True, data_only=False)
ws = wb['美团指标数据源']
header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))

# Build header index (last wins for duplicates)
header_idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

print('Header idx for key fields:')
for k in ['日期', '门店id', '门店编码', '品牌', '市场', '城市', '大区', '区域', '大店', '门店名称']:
    idx = header_idx.get(k, 'NOT FOUND')
    print(f'  {k}: {idx}')

# Read 5 rows
rows = list(ws.iter_rows(min_row=2, max_row=6, values_only=True))
wb.close()

for i, row in enumerate(rows):
    print(f'\nRow {i+2}:')
    code = str(row[header_idx.get('门店编码', 0)]).strip() if '门店编码' in header_idx else ''
    brand = str(row[header_idx.get('品牌', 0)]).strip() if '品牌' in header_idx else ''
    name_arch = str(row[header_idx.get('门店名称', 0)]).strip() if '门店名称' in header_idx else ''
    name_orig = str(row[1]).strip() if len(row) > 1 else ''
    print(f'  code={code}, brand={brand}, name_arch={name_arch}, name_orig={name_orig}')
